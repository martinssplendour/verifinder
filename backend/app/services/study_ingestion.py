from __future__ import annotations

import hashlib
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.services.dataset_utils import count_csv_rows, csv_header, csv_rows, sha256_file
from app.services.normalization import normalise_name


class StudySchemaError(ValueError):
    pass


@dataclass(frozen=True)
class StudentSponsorSnapshot:
    path: Path
    file_hash: str
    source_row_count: int
    record_count: int


@dataclass(frozen=True)
class OfsSnapshot:
    path: Path
    file_hash: str
    record_count: int


STUDENT_REQUIRED = {"Sponsor Name", "Sponsor Type", "Status", "Route"}


def inspect_student_sponsors(path: Path) -> StudentSponsorSnapshot:
    if not path.is_file():
        raise FileNotFoundError(path)
    missing = sorted(STUDENT_REQUIRED - set(csv_header(path)))
    if missing:
        raise StudySchemaError(f"Missing required student-sponsor columns: {missing}")
    source_row_count = count_csv_rows(path)
    record_count = sum(1 for _ in student_sponsor_rows(path, "inspection"))
    if record_count == 0:
        raise StudySchemaError("The student-sponsor CSV contains no usable providers.")
    return StudentSponsorSnapshot(path, sha256_file(path), source_row_count, record_count)


def student_sponsor_rows(path: Path, version_id: str):
    grouped: dict[str, dict] = {}
    for row in csv_rows(path):
        name = row.get("Sponsor Name", "").strip()
        if not name:
            continue
        town = row.get("Town/City", "").strip()
        sponsor_type = row.get("Sponsor Type", "").strip()
        status = row.get("Status", "").strip()
        identity = "|".join((normalise_name(name), normalise_name(town), normalise_name(sponsor_type), normalise_name(status)))
        record = grouped.setdefault(
            identity,
            {
                "dataset_version_id": version_id,
                "source_record_key": hashlib.sha1(identity.encode("utf-8")).hexdigest(),
                "organisation_name": name,
                "normalised_name": normalise_name(name),
                "town_city": town or None,
                "additional_locations": None,
                "sponsor_type": sponsor_type or None,
                "status": status or None,
                "routes": [],
                "immigration_compliance": None,
            },
        )
        route = row.get("Route", "").strip()
        if route and route not in record["routes"]:
            record["routes"].append(route)
        location = row.get("Additional Locations", "").strip()
        if location:
            record["additional_locations"] = location
        compliance = row.get("Immigration Compliance", "").strip()
        if compliance:
            record["immigration_compliance"] = compliance
    yield from grouped.values()


def _header_key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _find_header(headers: dict[str, int], *needles: str) -> int | None:
    return next((index for key, index in headers.items() if all(needle in key for needle in needles)), None)


def inspect_ofs_register(path: Path) -> OfsSnapshot:
    if not path.is_file():
        raise FileNotFoundError(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook["Register"]
        header_values = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
        keys = {_header_key(value) for value in header_values}
        if not any("ukprn" in key for key in keys) or not any("legal name" in key for key in keys):
            raise StudySchemaError("The OfS Register sheet does not contain its expected provider columns.")
        record_count = sum(1 for row in worksheet.iter_rows(min_row=4, values_only=True) if row[0] and row[3])
        workbook.close()
    except (KeyError, OSError, ValueError) as error:
        raise StudySchemaError(f"Could not read the OfS Register workbook: {error}") from error
    if record_count == 0:
        raise StudySchemaError("The OfS Register contains no active providers.")
    return OfsSnapshot(path, sha256_file(path), record_count)


def _text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _yes_no(value: object) -> bool | None:
    text = _text(value).lower()
    if text == "yes":
        return True
    if text == "no":
        return False
    return None


def _postcode(address: str) -> str | None:
    match = re.search(r"\b([A-Z]{1,2}\d[A-Z\d]?\s*\d[A-Z]{2})\b", address.upper())
    return match.group(1) if match else None


def ofs_provider_rows(snapshot: OfsSnapshot, version_id: str):
    workbook = load_workbook(snapshot.path, read_only=True, data_only=True)
    worksheet = workbook["Register"]
    header_row = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    headers = {_header_key(value): index for index, value in enumerate(header_row) if value}
    legal_index = _find_header(headers, "legal name")
    trading_index = _find_header(headers, "trading name")
    ukprn_index = _find_header(headers, "ukprn")
    address_index = _find_header(headers, "contact address")
    email_index = _find_header(headers, "email address")
    website_index = _find_header(headers, "website")
    charity_index = _find_header(headers, "charity")
    category_index = _find_header(headers, "category", "registered")
    fee_index = _find_header(headers, "fee limits")
    tef_index = _find_header(headers, "tef rating")
    powers_index = _find_header(headers, "highest level", "degree awarding powers")
    powers_date_index = _find_header(headers, "date of award", "degree awarding powers")
    university_index = _find_header(headers, "right to use", "university")
    university_date_index = _find_header(headers, "date", "university", "granted")
    university_basis_index = _find_header(headers, "how", "right to use", "university")
    access_pairs = [
        (index, index + 1)
        for key, index in headers.items()
        if key.startswith("is an access and participation plan in place") and index + 1 < len(header_row)
    ]
    condition_indexes = [index for key, index in headers.items() if "specific condition of registration imposed" in key]
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        if legal_index is None or ukprn_index is None:
            break
        legal_name = _text(row[legal_index])
        ukprn = _text(row[ukprn_index]).removesuffix(".0")
        if not legal_name or not ukprn:
            continue
        trading_raw = _text(row[trading_index]) if trading_index is not None else ""
        trading_names = [
            value.strip()
            for value in re.split(r"[\r\n]+", trading_raw)
            if value.strip() and value.strip().lower() != "not applicable"
        ]
        aliases = " | ".join(dict.fromkeys([normalise_name(legal_name), *(normalise_name(value) for value in trading_names)]))
        conditions = []
        for index in condition_indexes:
            value = _text(row[index])
            if value and value.lower() not in {"none", "not applicable"} and value not in conditions:
                conditions.append(value)
        access_plan = False
        access_url = None
        for status_index, url_index in access_pairs:
            if _yes_no(row[status_index]):
                access_plan = True
                candidate = _text(row[url_index])
                if candidate:
                    access_url = candidate
        address = _text(row[address_index]) if address_index is not None else ""
        yield {
            "dataset_version_id": version_id,
            "ukprn": ukprn,
            "legal_name": legal_name,
            "normalised_name": normalise_name(legal_name),
            "trading_names": trading_names,
            "normalised_aliases": aliases,
            "contact_address": address or None,
            "postcode": _postcode(address),
            "email": _text(row[email_index]) or None if email_index is not None else None,
            "website": _text(row[website_index]) or None if website_index is not None else None,
            "charity_status": _text(row[charity_index]) or None if charity_index is not None else None,
            "registration_category": _text(row[category_index]) or None if category_index is not None else None,
            "fee_limits": _text(row[fee_index]) or None if fee_index is not None else None,
            "tef_rating": _text(row[tef_index]) or None if tef_index is not None else None,
            "degree_awarding_powers": _text(row[powers_index]) or None if powers_index is not None else None,
            "degree_awarding_powers_date": _text(row[powers_date_index]) or None if powers_date_index is not None else None,
            "university_title": _yes_no(row[university_index]) if university_index is not None else None,
            "university_title_date": _text(row[university_date_index]) or None if university_date_index is not None else None,
            "university_title_basis": _text(row[university_basis_index]) or None if university_basis_index is not None else None,
            "access_plan": access_plan,
            "access_plan_url": access_url,
            "specific_conditions": conditions,
        }
    workbook.close()
